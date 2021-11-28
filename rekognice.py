from tkinter import *
from PIL import Image, ImageTk
from tkinter import messagebox
from tkinter import filedialog

import pyttsx3

import pickle

import face_recognition as fr
import cv2
import numpy as np

import os

from random import choice

from datetime import datetime, date

import time

from openpyxl import load_workbook
from openpyxl import Workbook

root = Tk()
root.title("ЯⲈKOGNICⲈ")
root.iconbitmap("logo.ico")

w = root.winfo_screenwidth()
h = root.winfo_screenheight()
root.geometry("%dx%d+0+0" % (w, h))
root.resizable(0, 0)


image = Image.open("bg.jpg")
image = image.resize((1550, 1000))
photo = ImageTk.PhotoImage(image)

l_image = Label(root, image=photo)
l_image.place(relx=0.5, rely=0.5, anchor="center")

f_home = Frame(l_image, height=500, width=800, bg='#191970')
f_home.place(relx=0.5, rely=0.5, anchor="center")

f_login = Frame(l_image, height=500, width=800, bg='#191970')
f_login.place(relx=0.5, rely=0.5, anchor="center")

f_register_t = Frame(l_image, height=500, width=800, bg='#191970')
f_register_t.place(relx=0.5, rely=0.5, anchor="center")

f_register_s = Frame(l_image, height=500, width=800, bg='#191970')
f_register_s.place(relx=0.5, rely=0.5, anchor="center")

f_student = Frame(l_image, height=500, width=800, bg='#191970')
f_student.place(relx=0.5, rely=0.5, anchor="center")

f_attendance = Frame(l_image, height=500, width=800, bg='#191970')
f_attendance.place(relx=0.5, rely=0.5, anchor="center")

f_excel = Frame(l_image, height=500, width=800, bg='#191970')
f_excel.place(relx=0.5, rely=0.5, anchor="center")

f_about = Frame(l_image, height=500, width=800, bg='#191970')
f_about.place(relx=0.5, rely=0.5, anchor="center")


uname_pswd_store = {}
uname_pswd_check = {}
student_details_store = {}
student_details_get = {}


user_image = Image.open("user.png")
user_image = user_image.resize((70, 70))
user_photo = ImageTk.PhotoImage(user_image)

pass_image = Image.open("pass.png")
pass_image = pass_image.resize((70, 70))
pass_photo = ImageTk.PhotoImage(pass_image)
 

# Login For Teachers

def login():
    global Username_entry, Password_entry
    f_login.tkraise()

    #LABELS

    label = Label(f_login, text="Login", font=('Comic Sans', 40,"bold"), bg="orange", fg="blue", width=35, pady=2)
    label.place(relx=0.5, rely=0.15, anchor='center')

    l_user_image = Label(f_login, image=user_photo)
    l_user_image.place(relx=0.30, rely=0.48, anchor="center")

    l_pass_image = Label(f_login, image=pass_photo)
    l_pass_image.place(relx=0.30, rely=0.65, anchor="center")

    #ENTRY BOXES

    Username_entry = Entry(f_login, bd=2, width=20, font=('Calibri', 24))
    Username_entry.place(relx=0.57, rely=0.48, anchor="center")

    Password_entry = Entry(f_login, bd=2, width=20,font=('Calibri', 24), show="*")
    Password_entry.place(relx=0.57, rely=0.65, anchor="center")


    #FUNCTIONS  
  
    def login_submit():
        global Username_entry, Password_entry
        u, p = Username_entry.get(), Password_entry.get()

        if len(u) == 0:
            messagebox.showinfo("ERROR", "Username cannot be left blank")

        elif len(p) == 0:
            messagebox.showinfo("ERROR", "Password cannot be left blank")

        with open("credentials.dat", "rb") as f:
            try:
                while True:
                    uname_pswd_check.update(pickle.load(f))
            except EOFError:
                pass

            if u in uname_pswd_check:    
                if uname_pswd_check[u] == p:
                    speak("Logged in Successfully")
                    teacher_menu.add_command(label="View", command=excel, state="normal")
                    f_excel.tkraise()
                    teacher_menu.entryconfig(0, state=DISABLED)
                    student_menu.entryconfig(0, state=DISABLED)
                else:
                    messagebox.showwarning("ERROR", "Please check your password ")
            else:
                messagebox.showwarning("ERROR", "No such user exists")

        Username_entry.delete(0, END)
        Password_entry.delete(0, END)


    #BUTTONS

    Submit_btn = Button(f_login, text="SUBMIT", font=("Comic Sans", 15, "bold"), bg="orange", fg="blue", command=login_submit)
    Submit_btn.place(relx=0.5, rely=0.85, anchor="center")


# Registration For Teachers

def T_Register():
    global uname_pswd_store, name1_entry, name2_entry, user_entry, passw1_entry, passw2_entry
    f_register_t.tkraise()

    #LABELS

    label = Label(f_register_t, text="Registration Form!", font=('Comic Sans', 30, "bold"), bg="orange", fg="blue", width=35, pady=2)
    label.place(relx=0.5, rely=0.1, anchor="center")

    name1_label = Label(f_register_t, text="Firstname  :", font=('Comic Sans', 13, "bold"), bg="orange", fg="blue")
    name1_label.place(relx=0.37, rely=0.25, anchor='center')

    name2_label = Label(f_register_t, text="Lastname  :", font=('Comic Sans', 13, "bold"), bg="orange", fg="blue")
    name2_label.place(relx=0.37, rely=0.38, anchor='center')

    user_label = Label(f_register_t, text=" Create your Username:", font=('Comic Sans', 13, "bold"), bg="orange", fg="blue")
    user_label.place(relx=0.31, rely=0.5, anchor='center')

    passw1_label = Label(f_register_t, text=" Create your Password:", font=('Comic Sans', 13, "bold"), bg="orange", fg="blue")
    passw1_label.place(relx=0.31, rely=0.62, anchor='center')

    passw2_label = Label(f_register_t, text="Re-enter your Password:",font=('Comic Sans', 13, "bold"), bg="orange", fg="blue")
    passw2_label.place(relx=0.31, rely=0.73, anchor='center')

    photo_label = Label(f_register_t, text="Passport size photo:", font=('Comic Sans', 13, "bold"), bg="orange", fg="blue")
    photo_label.place(relx=0.29, rely=0.845, anchor='center')

    #ENTRYBOXES

    name1_entry = Entry(f_register_t, bd=2, width=20, font=('Calibri', 14))
    name1_entry.place(relx=0.57, rely=0.25, anchor="center")

    name2_entry = Entry(f_register_t, bd=2, width=20, font=('Calibri', 14))
    name2_entry.place(relx=0.57, rely=0.37, anchor="center")

    user_entry = Entry(f_register_t, bd=2, width=20, font=('Calibri', 14))
    user_entry.place(relx=0.57, rely=0.495, anchor="center")

    passw1_entry = Entry(f_register_t, bd=2, width=20, font=('Calibri', 14), show="*")
    passw1_entry.place(relx=0.57, rely=0.615, anchor="center")

    passw2_entry = Entry(f_register_t, bd=2, width=20,font=('Calibri', 14), show="*")
    passw2_entry.place(relx=0.57, rely=0.73, anchor="center")


    #FUNCTIONS

    def rename(path, new_name, extension):
        list = os.listdir(path)
        os.chdir(path)
        try:
            for i in list:
                if i == "name.jpg":
                    os.rename(i, new_name+'.'+extension)
        except:
            print('error')


    def pho():
        global my_image, my_image_label, name1_entry, name2_entry, pho
        root.filename = filedialog.askopenfilename(title="select a file", filetypes=(("all files", "*.*"), ("jpg files", "*.jpg")))
        my_image = Image.open(root.filename)
        ###### Add path according to your system.    #######
        my_image = my_image.save(r"face_images\name.jpg")
        speak("Hurray! Your photo is uploaded")
        pho = True
        ###### Add path according to your system.    #######
        path = "\face_images"
        a = name1_entry.get() + ' '+name2_entry.get()
        rename(path, a, 'jpg')


    def register_submit():
        global uname_pswd_store, name1_entry, name2_entry, user_entry, passw1_entry
        from tkinter import messagebox
        a = passw1_entry.get()
        b = passw2_entry.get()
        c = user_entry.get()
        d = name1_entry.get()
        if len(name1_entry.get()) == 0:
            messagebox.showinfo("ERROR", "FirstName required")
        elif len(user_entry.get()) == 0:
            messagebox.showwarning("ERROR", "Username is not typed")
        elif len(passw1_entry.get()) == 0 and len(passw2_entry.get()) == 0:
            messagebox.showerror("ERROR", "Password is not typed")
        elif len(passw1_entry.get()) < 8 and len(passw2_entry.get()) < 8:
            messagebox.showwarning("ERROR", "password less than 8 charaters")
        elif passw1_entry.get() != passw2_entry.get():
            messagebox.showwarning("ERROR", "Passwords donot Match")
        elif pho != True:
            messagebox.showwarning("ERROR", "You need to upload you photo")
        else:
            if a == b:
                speak("Registered successfully")
                uname_pswd_store.update({user_entry.get(): passw1_entry.get()})
                with open("credentials.dat", 'ab') as f:
                    pickle.dump(uname_pswd_store, f)

                name1_entry.delete(0, END)
                name2_entry.delete(0, END)
                user_entry.delete(0, END)
                passw1_entry.delete(0, END)
                passw2_entry.delete(0, END)

            else:
                messagebox.showerror("ERROR", "please check your password")


    #BUTTONS

    photo_btn = Button(f_register_t, text="Upload Photo", font=("Comic Sans", 10, "bold"), bg="orange", fg="blue", command=pho)
    photo_btn.place(relx=0.544, rely=0.845, anchor="center")

    register_btn = Button(f_register_t, text="Register", font=("Comic Sans", 12, "bold"), bg="orange", fg="blue", command=register_submit)
    register_btn.place(relx=0.85, rely=0.9, anchor="center")


# Registration For Students

def S_Register():
    global name1_entry, name2_entry, usn_entry, class_sec_reg, student_details_store
    f_register_s.tkraise()

    #LABELS

    label = Label(f_register_s, text="Registration Form!", font=('Comic Sans', 30, "bold"), bg="orange", fg="blue", width=35, pady=2)
    label.place(relx=0.5, rely=0.1, anchor="center")

    name1_label = Label(f_register_s, text="Firstname :", font=('Comic Sans', 16, "bold"), bg="orange", fg="blue")
    name1_label.place(relx=0.32, rely=0.20, anchor='center')

    name2_label = Label(f_register_s, text="Lastname :", font=('Comic Sans', 16, "bold"), bg="orange", fg="blue")
    name2_label.place(relx=0.32, rely=0.35, anchor='center')

    photo_label = Label(f_register_s, text="Passport size photo:", font=('Comic Sans', 16, "bold"), bg="orange", fg="blue")
    photo_label.place(relx=0.29, rely=0.845, anchor='center')

    class_sec_label = Label(f_register_s, text="Class & Section:", font=('Comic Sans', 16, "bold"), bg="orange", fg="blue")
    class_sec_label.place(relx=0.29, rely=0.50, anchor='center')

    usn_label = Label(f_register_s, text="USN :",font=('Comic Sans', 16, "bold"), bg="orange", fg="blue")
    usn_label.place(relx=0.36, rely=0.65, anchor='center')

    #ENTRYBOXES

    name1_entry = Entry(f_register_s, bd=2, width=20, font=('Calibri', 18))
    name1_entry.place(relx=0.58, rely=0.20, anchor="center")

    name2_entry = Entry(f_register_s, bd=2, width=20, font=('Calibri', 18))
    name2_entry.place(relx=0.58, rely=0.35, anchor="center")
    
    usn_entry = Entry(f_register_s, bd=2, width=20, font=('Calibri', 18))
    usn_entry.place(relx=0.58, rely=0.65, anchor="center")

    classes_reg = [
        "XII F2",
        "XII F1",
        "XII E1",
        "XII E2",
        "XII E3",
        "XII G1",
        "XII G2",

    ]

    class_sec_reg = StringVar()
    class_sec_reg.set("Classes")

    drop = OptionMenu(f_register_s, class_sec_reg, *classes_reg)
    drop.place(relx=0.55, rely=0.50, anchor="center")
    drop.config(font=("Comic Sans", 16, "bold"), bg="orange", fg="blue")


    #FUNCTIONS
    def rename(path, new_name, extension):
        list = os.listdir(path)
        os.chdir(path)
        try:
            for i in list:
                if i == "name.jpg":
                    os.rename(i, new_name+'.'+extension)
        except:
            print('error')


    def pho():
        global my_image, my_image_label, photo
        root.filename = filedialog.askopenfilename(title="select a file", filetypes=(("all files", "*.*"), ("jpg files", "*.jpg")))
        my_image = Image.open(root.filename)
        ###### Add path according to your system.    #######
        my_image = my_image.save("face_images\name.jpg")
        speak("Hurray! your photo is uploaded")
        photo = True
        ###### Add path according to your system.    #######
        path = "\face_images"
        a = name1_entry.get() + ' ' + name2_entry.get()
        rename(path, a, 'jpg')

    def register_submit():
        global name1_entry, name2_entry, usn_entry, class_sec_reg, student_details_store
        from tkinter import messagebox
        student_check = {} 

        with open("student_details.dat", 'rb') as f_sd:
            try:
                while True:
                    student_check.update(pickle.load(f_sd))
            except EOFError:
                pass
            
            l = []
            for m, n in student_check.items():
                l.append(m)

            if len(name1_entry.get()) == 0:
                messagebox.showinfo("ERROR", "FirstName required")
            elif len(name2_entry.get()) == 0:
                messagebox.showinfo("ERROR", "LastName required")
            elif class_sec_reg.get()== "Classes":
                messagebox.showinfo("ERROR", "Class required")
            elif usn_entry.get() in l:
                speak("Student already Exists!")
                name1_entry.delete(0, END)
                name2_entry.delete(0, END)
                usn_entry.delete(0, END)
            elif photo != True:
                messagebox.showinfo("ERROR", "You need to upload your photo")
                return
            else:
                speak("Registered successfully")
                student_details_store.update({usn_entry.get(): [name1_entry.get()+" "+name2_entry.get(), class_sec_reg.get()]})
                with open("student_details.dat", 'ab') as f_sd:
                    pickle.dump(student_details_store, f_sd)

        name1_entry.delete(0, END)
        name2_entry.delete(0, END)
        usn_entry.delete(0, END)
        
    #BUTTONS

    photo_btn = Button(f_register_s, text="Upload Photo", font=("Comic Sans",15,"bold"),bg="orange",fg="blue",command=pho)
    photo_btn.place(relx=0.544, rely=0.845, anchor="center")

    register_btn = Button(f_register_s, text="Register", font=("Comic Sans",16,"bold"),bg="orange",fg="blue",command=register_submit)
    register_btn.place(relx=0.85, rely=0.9, anchor="center")


def back():
    f_home.tkraise()

def student_page():
    global usn_entry
    f_student.tkraise()

    heading_label = Label(f_student, text="Attendance", font=('Comic Sans',28,"bold"),bg="orange",fg="blue",width=35, pady=2)
    heading_label.place(relx=0.5, rely=0.12, anchor='center')

    usn_label = Label(f_student, text="Enter your USN :", font=('Comic Sans',18,"bold"),bg="orange",fg="blue")
    usn_label.place(relx=0.3, rely=0.25, anchor='center')

    usn_entry = Entry(f_student, bd=2, width=20,font=('Calibri',20))
    usn_entry.place(relx=0.68, rely=0.25, anchor="center")

    l_boy_image = Label(f_student, image=boy_photo)
    l_boy_image.place(relx=0.5, rely=0.6, anchor="center")

    go_btn = Button(f_student, text="GO", font=("Comic Sans",14,"bold"),bg="orange",fg="blue",command = attendance_page)
    go_btn.place(relx=0.5, rely=0.9, anchor="center")

    
def attendance_page():
    global full_name, class_section
    ######     Add path according to your system.    #######
    path = r"face_images"
    myList = os.listdir(path)

    with open("student_details.dat", 'rb') as f_sd:
        try:
            while True:
                student_details_get.update(pickle.load(f_sd))
        except EOFError:
            pass
        for m, n in student_details_get.items():
            if usn_entry.get() in m:
                full_name = n[0]
                class_section = n[1]
        
    try:
        for i in range(1):
            if full_name + '.jpg' in myList:
                student_menu.entryconfig(0, state=DISABLED)
                f_attendance.tkraise()

                hour = int(datetime.now().hour)
                if hour >= 0 and hour < 12:
                    speak("Good Morning" + full_name)

                elif hour >= 12 and hour < 18:
                    speak("Good Afternoon" + full_name)

                else:
                    speak("Good Evening" + full_name)

                def clock():
                    hour = time.strftime("%H")
                    minute = time.strftime("%M")
                    second = time.strftime("%S")
                    day = time.strftime("%A")
                    am_pm = time.strftime("%p")
                    date = time.strftime("%x")

                    l_time.config(text=hour + ":" + minute + ":" + second + " " + am_pm)
                    l_time.after(1000, clock)

                    l_day.config(text=day)

                    l_date.config(text=date)

                l_time = Label(f_attendance, text = "", font = ("Comic Sans", 36, "bold"), bg="#191970", fg="red2")
                l_time.place(relx=0.5, rely=0.2, anchor="center")

                l_day = Label(f_attendance, text = "", font = ("Comic Sans", 22, "bold"), bg="#191970", fg="red2")
                l_day.place(relx=0.65, rely=0.4, anchor="center")

                l_date = Label(f_attendance, text = "", font = ("Comic Sans", 22, "bold"), bg="#191970", fg="red2")
                l_date.place(relx=0.35, rely=0.4, anchor="center")

                clock()

                B_attendance = Button(f_attendance, text="Attendance", font=("Comic Sans",16,"bold"),bg="orange",fg="blue", command= attendance_check)
                B_attendance.place(relx=0.3, rely=0.8, anchor="center")

                B_finish = Button(f_attendance, text="Finish", font=("Comic Sans",16,"bold"),bg="orange",fg="blue", command = end)
                B_finish.place(relx=0.7, rely=0.8, anchor="center")

                with open("quotes.txt", 'r')as f:
                    a = f.readlines()
                    l = list(a)
                    quotes_p = choice(l)

                quotes = Label(f_attendance, text= quotes_p, font = ("Comic Sans MS", 14, "bold"), fg = "orange", bg ="#191970")
                quotes.place(relx=0.5, rely=0.6, anchor="center")

    except:
        speak("Oops! You Need Register First")
        exit()
        

def attendance_check():
    ###### Add path according to your system.    #######
    path = r"face_images"+"\\"+ full_name + ".jpg"

    images = []
    imagenames = []

    cimg = cv2.imread(path)
    images.append(cimg)
    imagenames.append(full_name)

    speak("Encoding Started")

    encodedimages = []
    cimg = cv2.cvtColor(cimg, cv2.COLOR_BGR2RGB)
    encode = fr.face_encodings(cimg)[0]
    encodedimages.append(encode)

    encodeListKnown = encodedimages
    speak("Encoding Completed")

    # Starting webcam
    video_capture = cv2.VideoCapture(0)

    video_capture.set(3, 1920)
    video_capture.set(4, 1080)

    photo_captured = False

    start_time = time.time()

    while not photo_captured and time.time() < start_time + 30:

        # Will take a single frame of the video
        ret, frame = video_capture.read()

        # To resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), None, fx=0.25, fy=0.25)

        # Convert the image from BGR color to RGB color
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        # To find all the faces and face encodings in the current frame of video
        face_locations = fr.face_locations(rgb_small_frame)
        face_encodings = fr.face_encodings(rgb_small_frame, face_locations)

        # To display the results
        face_names = []
        for faceencode, faceLoc in zip(face_encodings, face_locations):

            # To check if the current face is a match for the already known face(s)
            match = fr.compare_faces(encodeListKnown, faceencode)
            faceDis = fr.face_distance(encodeListKnown, faceencode)
            matchIndex = np.argmin(faceDis)

            if faceDis < 0.5:
                if match[matchIndex]:
                    name = imagenames[matchIndex].upper()
                    face_names.append(name)

                    for (top, right, bottom, left), name in zip(face_locations, face_names):

                        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
                        top *= 4
                        right *= 4
                        bottom *= 4
                        left *= 4

                        # For drawing a box around the face
                        cv2.rectangle(frame, (left, top),(right, bottom), (0, 0, 255), 2)

                        # For drawing a label with the name below the face
                        cv2.rectangle(frame, (left, bottom - 35),(right, bottom), (0, 0, 255), cv2.FILLED)
                        font = cv2.FONT_HERSHEY_DUPLEX

                        # For adding the name to the label
                        cv2.putText(frame, name, (left + 6, bottom - 6),font, 1.0, (255, 255, 255), 1)
                        mark_attendance(name)

        # To display the resulting image
        cv2.imshow('Video', frame)
        
        if cv2.waitKey(1) & len(face_names) > 0:
            photo_captured = True
            speak("Amigo! You are marked present")
        

    video_capture.release()
    cv2.destroyAllWindows()
    if photo_captured == False: 

        speak("Oops! You are not" + full_name)
        ###### Add path according to your system.    #######
        path = r"Excel\failed_attempts.xlsx"
        wb = load_workbook(path)

        sheet = wb.active
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15

        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Date"
        sheet.cell(row=1, column=3).value = "Time"

        current_row = sheet.max_row
        current_column = sheet.max_column

        now = datetime.now()
        timestring = now.strftime('%H:%M:%S')
        datestring = date.today()

        sheet.cell(row=current_row + 1, column=1).value = full_name
        sheet.cell(row=current_row + 1, column=2).value = datestring
        sheet.cell(row=current_row + 1, column=3).value = timestring

        ###### Add path according to your system.    #######
        wb.save(filename=r"Excel\failed_attempts.xlsx")
        exit(0)

###### Add path according to your system.    #######
boy_image = Image.open(r"inapp_images\boy.jpg")
boy_image = boy_image.resize((200, 240))
boy_photo = ImageTk.PhotoImage(boy_image)


def mark_attendance(name):

    ###### Add path according to your system.    #######
    path_main = r"\Excel"
    myList = os.listdir(path_main)

    if class_section + '.xlsx' in myList:
        ###### Add path according to your system.    #######
        path = r"Excel"+"\\"+class_section+".xlsx"
        wb = load_workbook(path)
    else:
        wb = Workbook()

    sheet = wb.active
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Date"
    sheet.cell(row=1, column=3).value = "Time"

    current_row = sheet.max_row
    current_column = sheet.max_column

    now = datetime.now()
    timestring = now.strftime('%H:%M:%S')
    datestring = date.today()

    sheet.cell(row=current_row + 1, column=1).value = name
    sheet.cell(row=current_row + 1, column=2).value = datestring
    sheet.cell(row=current_row + 1, column=3).value = timestring

    ###### Add path according to your system.    #######
    wb.save(filename=r"\Excel"+"\\"+class_section+".xlsx")

# Text-Speech 

engine = pyttsx3.init('sapi5')

# getting details of current voice
voices = engine.getProperty('voices')  

engine.setProperty('voice', voices[1].id)

def speak(audio):
    engine.say(audio)

    # Without this command, speech will not be audible to us.
    engine.runAndWait()


def open_excel():
    cl_sc = class_sec_excel.get()

    ###### Add path according to your system.    #######
    file = r"\Excel"+"\\" + cl_sc + ".xlsx"
    try:
        os.startfile(file)
    except FileNotFoundError:
        speak("No such class exists")


def open_failed_attempts():
    ###### Add path according to your system.    #######
    os.startfile(r"Excel\failed_attempts.xlsx")


def excel():
    f_excel.tkraise()


classes_excel = [
    "XII F2",
    "XII F1",
    "XII E1",
    "XII E2",
    "XII E3",
    "XII G1",
    "XII G2",

]

class_sec_excel = StringVar()
class_sec_excel.set("Classes")

drop = OptionMenu(f_excel, class_sec_excel, *classes_excel)
drop.place(relx=0.65, rely=0.6, anchor="center")
drop.config(font = ("Comic Sans", 16, "bold"), bg="orange",fg="blue")

###### Add path according to your system.    #######
class_image = Image.open(r"\inapp_images\class.jpg")
class_image = class_image.resize((140, 140))
class_photo = ImageTk.PhotoImage(class_image)

l_heading = Label(f_excel, text = "View Attendance!", font = ("Comic Sans", 28, "bold"),bg="orange",fg="blue",width=35,pady=2)
l_heading.place(relx=0.5, rely=0.155, anchor="center")

l_class = Label(f_excel, text = "Select a class", font = ("Comic Sans", 16, "bold"),bg="orange",fg="blue")
l_class.place(relx=0.4, rely=0.6, anchor="center")

l_class_image = Label(f_excel, image=class_photo)
l_class_image.place(relx=0.5, rely=0.38, anchor="center")

B_excel = Button(f_excel, text="Open Excel", font=("Comic Sans",12,"bold"),bg="orange",fg="blue", command= open_excel)
B_excel.place(relx=0.5, rely=0.71, anchor="center")

l_failed_attempts = Label(f_excel, text = "To view the failed attempts: ", font = ("Comic Sans", 16, "bold"),bg="orange",fg="blue")
l_failed_attempts.place(relx=0.4, rely=0.85, anchor="center")

B_failed_attempts = Button(f_excel, text="Click here!", font=("Comic Sans",12,"bold"),bg="orange",fg="blue", command= open_failed_attempts)
B_failed_attempts.place(relx=0.70, rely=0.85, anchor="center")


def end():
    try:
        speak("Have a good day" + full_name)
        quit()
    except NameError:
        speak("Have a good day")
        quit()

###### Add path according to your system.    #######
logo1_image = Image.open(r"inapp_images\logo.jpg")
logo1_image = logo1_image.resize((150, 150))
logo1_photo = ImageTk.PhotoImage(logo1_image)


def about_page():
    f_about.tkraise()
    label = Label(f_about, text="ЯⲈKOGNICⲈ", fg="orange", bg="#191970",font=("Algerian", 70, "bold"))
    label.place(relx=0.46, rely=0.17, anchor="center")

    l_logo1_image = Label(f_about, image=logo1_photo)
    l_logo1_image.place(relx=0.9, rely=0.17, anchor="center")

    info_label = Label(f_about, text='''
    This is an attendance app!
                        You need to be registered either as as student or teacher 
    for the app to recognise you
                The teacher can check the attendance directly from the app 
                by logging in with his/her own credentials
                    Have a Good Day!''', fg="orange", bg="#191970",font=("Comic Sans",18, "bold"))
    info_label.place(relx=0.44, rely=0.56, anchor="center")


# Home Screen
###### Add path according to your system.    #######
cam_image = Image.open(r"inapp_images\cam.jpg")
cam_image = cam_image.resize((250, 200))
cam_photo = ImageTk.PhotoImage(cam_image)

###### Add path according to your system.    #######
mac_image = Image.open(r"inapp_images\mac.jpg")
mac_image = mac_image.resize((250, 200))
mac_photo = ImageTk.PhotoImage(mac_image)

###### Add path according to your system.    #######
logo_image = Image.open(r"\inapp_images\logo.jpg")
logo_image = logo_image.resize((150, 150))
logo_photo = ImageTk.PhotoImage(logo_image)

l_cam_image = Label(f_home, image=cam_photo)
l_cam_image.place(relx=0.7, rely=0.65, anchor="center")

l_mac_image = Label(f_home, image=mac_photo)
l_mac_image.place(relx=0.3, rely=0.65, anchor="center")

l_logo_image = Label(f_home, image=logo_photo)
l_logo_image.place(relx=0.9, rely=0.2, anchor="center")

label = Label(f_home, text="ЯⲈKOGNICⲈ", fg="orange", bg="#191970",font=("Algerian", 70, "bold"))
label.place(relx=0.45, rely=0.2, anchor="center")

f_home.tkraise()

# Menu Settings

my_menu = Menu(root)

file_menu = Menu(my_menu, tearoff=0)
teacher_menu = Menu(my_menu, tearoff=0)
student_menu = Menu(my_menu, tearoff=0)
help_menu = Menu(my_menu, tearoff=0)
about_menu = Menu(my_menu, tearoff=0)

my_menu.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="Home", command=back)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=end)

my_menu.add_cascade(label="Teacher", menu=teacher_menu)
teacher_menu.add_command(label="Login", command=login)
teacher_menu.add_separator()
teacher_menu.add_command(label="Register", command=T_Register)

my_menu.add_cascade(label="Student", menu=student_menu)
student_menu.add_command(label="Attendance", command=student_page)
student_menu.add_separator()
student_menu.add_command(label="Register", command=S_Register)

my_menu.add_cascade(label="About", menu=about_menu)
about_menu.add_command(label="Info", command=about_page)

root.config(menu=my_menu)

root.mainloop()
